#!/usr/bin/env python3
"""
build_dataset.py — Ingest IMNU MBA Term-IV elective rosters + master schedule
workbook into a normalized SQLite DB + CSVs + a validation report.

Faculty keyed on EMAIL. Subjects keyed on master COURSE CODE. Abbreviations
reconciled via an explicit alias map with a collision guard (S&DM->SDM allowed,
I&PM->IPM refused). Master = source of truth for subject/faculty/room/time;
rosters = source of truth for enrollment. Full rebuild each run => idempotent.

If the master workbook is not found in --input, an embedded snapshot (captured
from the user's own schedule file) is used so the build still completes.
"""
import argparse, glob, os, re, sqlite3, csv
from datetime import datetime
from openpyxl import load_workbook

REPORT = []
def note(sev, msg): REPORT.append((sev, msg))

# ----------------------------------------------------------------- helpers
def norm_abbr(s):
    return None if s is None else re.sub(r'\s+', '', str(s)).upper()

SECTION_RE = re.compile(r"^\s*([A-Za-z&]+)\s*\(\s*'?\s*([A-Za-z])\s*\)\s*$")
def parse_section_token(raw):
    if raw is None: return None
    raw = str(raw).strip()
    if not raw: return None
    m = SECTION_RE.match(raw)
    if m: return norm_abbr(m.group(1)), m.group(2).upper()
    if re.fullmatch(r"[A-Za-z&]{2,6}", raw): return norm_abbr(raw), None
    return None

EMAIL_RE = re.compile(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}")
def extract_emails(s): return EMAIL_RE.findall(s or "")

def clean_faculty_name(s):
    if not s: return ""
    s = re.sub(r"\(.*?\)", " ", s)
    s = re.sub(r"\bProf\.?\b", " ", s, flags=re.I)
    s = re.sub(r"\bVF\b", " ", s, flags=re.I)
    s = re.sub(r"\bDr\.?\b", " ", s, flags=re.I)
    s = re.sub(r"\bM[rs]s?\.?\b", " ", s)
    return re.sub(r"\s+", " ", s).strip(" .,/")

def is_visiting(s): return bool(s) and bool(re.search(r"\bVF\b|visiting", s, re.I))
def batch_of(roll):
    m = re.match(r"\s*(\d{2}[A-Za-z]+)", str(roll))
    return m.group(1).upper() if m else None

def clean_time(s):
    if not s: return (None, None)
    parts = re.findall(r"\d{1,2}[.:]\d{2}\s*[AP]M", str(s), re.I)
    parts = [re.sub(r"\s+", "", p).upper().replace(".", ":") for p in parts]
    if len(parts) >= 2: return (parts[0], parts[1])
    if len(parts) == 1: return (parts[0], None)
    return (None, None)

DIV_SPEC_RE = re.compile(r"\(\s*Div\.?\s*([^)]*)\)", re.I)
def parse_faculty_cell(text):
    text = text or ""
    matches = list(DIV_SPEC_RE.finditer(text))
    groups = []
    if matches:
        start = 0
        for m in matches:
            seg = text[start:m.start()]
            divs = [d.upper() for d in re.findall(r"[A-Fa-f]", m.group(1))]
            groups.append((seg, divs)); start = m.end()
        tail = text[start:].strip()
        if tail and groups:
            groups[-1] = (groups[-1][0] + " " + tail, groups[-1][1])
    else:
        groups = [(text, None)]
    out = []
    for seg, divs in groups:
        names = [clean_faculty_name(n) for n in re.split(r"[/&\n]", seg)]
        names = [n for n in names if n]
        if names: out.append({"names": names, "divisions": divs, "raw": seg})
    return out

# ----------------------------------------------------------------- master
def _build_master(cd_rows, grid_rows, source):
    subjects, faculty, classrooms = {}, {}, set()
    division_map, alias_map = {}, {}
    for r in cd_rows[1:]:
        code = r[1] if len(r) > 1 else None
        if code is None or (isinstance(code, str) and "total" in code.lower()):
            continue
        code = str(code).strip()
        abbr = str(r[2]).strip() if len(r) > 2 and r[2] else None
        name = str(r[3]).strip() if len(r) > 3 and r[3] else None
        area = str(r[4]).strip() if len(r) > 4 and r[4] else None
        credits = r[5] if len(r) > 5 else None
        fac_cell = r[6] if len(r) > 6 else None
        room_cell = r[7] if len(r) > 7 else None
        email_cell = r[8] if len(r) > 8 else None
        subjects[code] = {"code": code, "abbr": abbr, "name": name, "area": area, "credits": credits}
        na = norm_abbr(abbr)
        if na:
            alias_map.setdefault(na, code)
            amp = na.replace("&", "")
            if amp != na and amp not in alias_map:
                alias_map[amp] = code
            elif amp != na and alias_map.get(amp) != code:
                note("INFO", f"Alias '{amp}' kept distinct from {code} ({abbr}) "
                             f"(already maps to {alias_map[amp]}) — e.g. IPM vs I&PM.")
        groups = parse_faculty_cell(fac_cell)
        room_lines = [x.strip() for x in str(room_cell or "").split("\n") if x.strip()]
        emails = extract_emails(str(email_cell or ""))
        flat = []
        for gi, g in enumerate(groups):
            room = room_lines[gi] if gi < len(room_lines) else (room_lines[-1] if room_lines else None)
            for nm in g["names"]:
                flat.append({"name": nm, "divisions": g["divisions"], "room": room,
                             "visiting": is_visiting(g["raw"])})
        if len(emails) < len(flat):
            note("WARN", f"{code} ({abbr}): {len(flat)} faculty but {len(emails)} email(s); "
                         f"some faculty lack an email key.")
        for i, fac in enumerate(flat):
            email = emails[i] if i < len(emails) else None
            key = email or f"noemail::{code}::{fac['name']}"
            faculty.setdefault(key, {"email": email, "name": fac["name"],
                                     "is_visiting": int(fac["visiting"])})
            if fac["room"]: classrooms.add(fac["room"])
            for d in (fac["divisions"] if fac["divisions"] else [None]):
                division_map[(code, d)] = {"faculty_key": key, "room": fac["room"]}

    meetings = []
    events = []
    if grid_rows:
        header, times = grid_rows[0], grid_rows[1]
        session_cols = []
        for c in range(min(13, len(header))):
            v = header[c]
            if isinstance(v, str) and v.strip().lower().startswith("session"):
                t = times[c] if c < len(times) else None
                session_cols.append((c, v.strip(), clean_time(t)))
        for r in grid_rows[3:]:
            if not r or not isinstance(r[0], datetime) or r[0].year != 2026:
                continue
            date, day = r[0], (str(r[1]).strip() if len(r) > 1 and r[1] else None)
            seen_ev = set()
            for c in range(2, min(13, len(r))):
                cell = r[c]
                if not isinstance(cell, str): continue
                low = cell.lower()
                if "holiday" in low:
                    nm = re.sub(r"\(?\s*holiday\s*\)?", "", cell, flags=re.I).strip(" -\u2013\u2014()")
                    if ("h", nm) not in seen_ev:
                        events.append({"date": date.strftime("%Y-%m-%d"), "day": day,
                                       "type": "holiday", "name": nm or "Holiday"})
                        seen_ev.add(("h", nm))
                elif re.search(r"\bexam", low):
                    if ("e", cell) not in seen_ev:
                        events.append({"date": date.strftime("%Y-%m-%d"), "day": day,
                                       "type": "exam", "name": cell.strip()})
                        seen_ev.add(("e", cell))
            for c, sess, (t0, t1) in session_cols:
                cell = r[c] if c < len(r) else None
                if not isinstance(cell, str): continue
                for tok in cell.split("/"):
                    parsed = parse_section_token(tok)
                    if not parsed: continue
                    ab, div = parsed
                    if ab not in alias_map: continue
                    meetings.append({"abbr_norm": ab, "division": div,
                                     "date": date.strftime("%Y-%m-%d"), "day": day,
                                     "session": sess, "start": t0, "end": t1})
    note("INFO", f"Master catalog source: {source}.")
    return subjects, faculty, classrooms, division_map, alias_map, meetings, events

def find_master(input_dir):
    cands = [f for f in glob.glob(os.path.join(input_dir, "*.xlsx"))
             if any(k in os.path.basename(f).lower() for k in ("schedule", "course detail", "term"))]
    return max(cands, key=os.path.getsize) if cands else None

def parse_master(input_dir):
    path = find_master(input_dir)
    if path:
        wb = load_workbook(path, read_only=True, data_only=True)
        cd = next((s for s in wb.sheetnames if "course detail" in s.lower()), wb.sheetnames[0])
        cd_rows = [list(r) for r in wb[cd].iter_rows(values_only=True)]
        gn = next((s for s in wb.sheetnames if re.search(r"\d{2}\.\d{2}\.\d{4}", s)), None)
        grid_rows = [list(r) for r in wb[gn].iter_rows(values_only=True)] if gn else []
        wb.close()
        return _build_master(cd_rows, grid_rows, os.path.basename(path))
    note("WARN", "Master schedule workbook not found in input folder; used the embedded "
                 "snapshot captured from your schedule file. Re-add that .xlsx to refresh "
                 "subjects, faculty emails, classrooms, and the weekly timetable.")
    return _build_master(EMBEDDED_CD, EMBEDDED_GRID, "embedded snapshot")

# ----------------------------------------------------------------- rosters
def parse_roster(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    results = []
    for sn in wb.sheetnames:
        rows = [list(r) for r in wb[sn].iter_rows(values_only=True)]
        if not rows: continue
        hidx = None
        for i, r in enumerate(rows[:6]):
            if any(isinstance(c, str) and c.strip().lower().startswith("roll no") for c in r):
                hidx = i; break
        if hidx is None:
            note("WARN", f"{os.path.basename(path)}/'{sn}': no 'Roll No.' header; skipped."); continue
        header = rows[hidx]
        roll_cols = [j for j, c in enumerate(header)
                     if isinstance(c, str) and c.strip().lower().startswith("roll no")]
        for j in roll_cols:
            code_cell = rows[0][j] if j < len(rows[0]) else None
            fac_cell = rows[0][j+1] if j+1 < len(rows[0]) else None
            parsed = parse_section_token(code_cell)
            if not parsed:
                note("WARN", f"{os.path.basename(path)}/'{sn}': bad section header "
                             f"'{code_cell}'; skipped."); continue
            ab, div = parsed
            students = []
            for r in rows[hidx+1:]:
                roll = r[j] if j < len(r) else None
                name = r[j+1] if j+1 < len(r) else None
                if roll is None: continue
                if isinstance(roll, str) and roll.strip().lower().startswith("roll no"): continue
                students.append((str(roll).strip(), str(name).strip() if name else None))
            results.append({"abbr_norm": ab, "division": div,
                            "faculty_raw": str(fac_cell).strip() if fac_cell else None,
                            "students": students,
                            "source": f"{os.path.basename(path)}::{sn}"})
    wb.close()
    return results

# ----------------------------------------------------------------- build
def build(input_dir, output_dir):
    os.makedirs(output_dir, exist_ok=True)
    csv_dir = os.path.join(output_dir, "csv"); os.makedirs(csv_dir, exist_ok=True)
    master_path = find_master(input_dir)
    subjects, faculty, classrooms, division_map, alias_map, meetings, events = parse_master(input_dir)
    roster_files = sorted(f for f in glob.glob(os.path.join(input_dir, "*.xlsx")) if f != master_path)

    students, name_variants, enrollments, sections = {}, {}, [], {}
    for rf in roster_files:
        for sec in parse_roster(rf):
            ab, div = sec["abbr_norm"], sec["division"]
            code = alias_map.get(ab)
            if not code:
                note("ERROR", f"Roster abbreviation '{ab}' ({os.path.basename(rf)}) not in master "
                              f"catalog; section dropped."); continue
            dm = division_map.get((code, div)) or division_map.get((code, None))
            if dm is None:
                note("WARN", f"No master faculty/room for {subjects[code]['abbr']}({div}); using roster "
                             f"name '{sec['faculty_raw']}'.")
                fk = f"noemail::{code}::{clean_faculty_name(sec['faculty_raw'])}"
                faculty.setdefault(fk, {"email": None, "name": clean_faculty_name(sec["faculty_raw"]),
                                        "is_visiting": int(is_visiting(sec["faculty_raw"]))})
                room = None
            else:
                fk, room = dm["faculty_key"], dm["room"]
                if sec["faculty_raw"]:
                    rn = clean_faculty_name(sec["faculty_raw"]); mn = faculty[fk]["name"]
                    if rn and mn and rn.lower() != mn.lower() and rn.split()[-1].lower() != mn.split()[-1].lower():
                        note("INFO", f"Faculty name differs for {subjects[code]['abbr']}({div}): "
                                     f"roster '{rn}' vs master '{mn}'.")
            sections.setdefault((code, div), {"faculty_key": fk, "room": room, "sources": set()})
            sections[(code, div)]["sources"].add(sec["source"])
            for roll, name in sec["students"]:
                if roll not in students: students[roll] = name
                elif name and students[roll] and name != students[roll]:
                    name_variants.setdefault(roll, set()).add(name)
                enrollments.append((roll, code, div))

    for m in meetings:
        code = alias_map.get(m["abbr_norm"]); 
        if not code: continue
        div = m["division"]
        if (code, div) not in sections:
            dm = division_map.get((code, div)) or division_map.get((code, None))
            if dm is None: continue
            sections[(code, div)] = {"faculty_key": dm["faculty_key"], "room": dm["room"],
                                     "sources": {"timetable-only"}}
            note("WARN", f"Section {subjects[code]['abbr']}({div}) is in the timetable but has no "
                         f"roster file (0 enrolled students).")

    db_path = os.path.join(output_dir, "university.db")
    if os.path.exists(db_path): os.remove(db_path)
    con = sqlite3.connect(db_path); cur = con.cursor()
    cur.executescript("""
    CREATE TABLE subjects(code TEXT PRIMARY KEY, abbr TEXT, name TEXT, area TEXT, credits REAL);
    CREATE TABLE faculty(faculty_key TEXT PRIMARY KEY, email TEXT, name TEXT, is_visiting INTEGER);
    CREATE TABLE classrooms(code TEXT PRIMARY KEY);
    CREATE TABLE sections(section_id INTEGER PRIMARY KEY AUTOINCREMENT, subject_code TEXT,
        division TEXT, faculty_key TEXT, classroom_code TEXT, UNIQUE(subject_code, division));
    CREATE TABLE students(roll_no TEXT PRIMARY KEY, name TEXT, batch TEXT);
    CREATE TABLE enrollments(roll_no TEXT, section_id INTEGER, PRIMARY KEY(roll_no, section_id));
    CREATE TABLE meetings(section_id INTEGER, week_of TEXT, date TEXT, day TEXT,
        session TEXT, start_time TEXT, end_time TEXT);
    CREATE TABLE events(date TEXT, day TEXT, type TEXT, name TEXT);
    """)
    for s in subjects.values():
        cur.execute("INSERT INTO subjects VALUES (?,?,?,?,?)",
                    (s["code"], s["abbr"], s["name"], s["area"], s["credits"]))
    for k, f in faculty.items():
        cur.execute("INSERT OR IGNORE INTO faculty VALUES (?,?,?,?)",
                    (k, f.get("email"), f["name"], f["is_visiting"]))
    for c in sorted(classrooms):
        cur.execute("INSERT OR IGNORE INTO classrooms VALUES (?)", (c,))
    sec_id = {}
    for (code, div), meta in sorted(sections.items()):
        room = meta["room"]
        if room and room not in classrooms:
            cur.execute("INSERT OR IGNORE INTO classrooms VALUES (?)", (room,)); classrooms.add(room)
        cur.execute("INSERT INTO sections(subject_code,division,faculty_key,classroom_code) VALUES (?,?,?,?)",
                    (code, div, meta["faculty_key"], room))
        sec_id[(code, div)] = cur.lastrowid
    for roll, name in students.items():
        cur.execute("INSERT OR IGNORE INTO students VALUES (?,?,?)", (roll, name, batch_of(roll)))
    seen = set()
    for roll, code, div in enrollments:
        sid = sec_id.get((code, div))
        if sid and (roll, sid) not in seen:
            cur.execute("INSERT OR IGNORE INTO enrollments VALUES (?,?)", (roll, sid)); seen.add((roll, sid))
    for m in meetings:
        code = alias_map.get(m["abbr_norm"])
        sid = sec_id.get((code, m["division"])) or sec_id.get((code, None))
        if sid:
            cur.execute("INSERT INTO meetings VALUES (?,?,?,?,?,?,?)",
                        (sid, "2026-06-22", m["date"], m["day"], m["session"], m["start"], m["end"]))
    for ev in events:
        cur.execute("INSERT INTO events VALUES (?,?,?,?)", (ev["date"], ev["day"], ev["type"], ev["name"]))
    con.commit()
    for tbl in ["subjects","faculty","classrooms","sections","students","enrollments","meetings","events"]:
        cur.execute(f"SELECT * FROM {tbl}"); cols = [d[0] for d in cur.description]
        with open(os.path.join(csv_dir, f"{tbl}.csv"), "w", newline="") as fh:
            w = csv.writer(fh); w.writerow(cols); w.writerows(cur.fetchall())
    for roll, variants in name_variants.items():
        note("INFO", f"Roll {roll}: multiple name spellings — {students[roll]!r} / {', '.join(sorted(variants))}.")
    counts = {t: cur.execute(f"SELECT COUNT(*) FROM {t}").fetchone()[0]
              for t in ["subjects","faculty","classrooms","sections","students","enrollments","meetings"]}
    con.close()
    write_report(output_dir, counts, master_path, roster_files)
    return counts

def write_report(output_dir, counts, master_path, roster_files):
    order = {"ERROR":0,"WARN":1,"INFO":2}
    rep = sorted(REPORT, key=lambda x: order.get(x[0],3))
    lines = ["# Import validation report","",f"Generated: {datetime.now():%Y-%m-%d %H:%M}","",
             f"Master workbook: {os.path.basename(master_path) if master_path else 'NOT in upload (embedded snapshot used)'}",
             f"Roster files: {len(roster_files)}","","## Row counts",""]
    for t,n in counts.items(): lines.append(f"- **{t}**: {n}")
    lines += ["","## Issues flagged for review",""]
    if not rep: lines.append("_None — everything reconciled cleanly._")
    for sev,msg in rep: lines.append(f"- **{sev}** — {msg}")
    with open(os.path.join(output_dir,"validation_report.md"),"w") as fh:
        fh.write("\n".join(lines)+"\n")

# ----------------------------------------------------------------- embedded snapshot
_H = [None]*11
EMBEDDED_CD = [
 _H,
 [1,'7MP117SE26','IPM','Investment and Portfolio Management','Finance',3,'Prof. Nikunj Patel (Div. A&B)\nProf. Bhavesh Patel (Div. C)\nProf. Dipti Saraf (Div. D)','T6\nT5\n309-F','nikunj@nirmauni.ac.in,\nbhavesh@nirmauni.ac.in,\ndipti.saraf@nirmauni.ac.in','2\n1\n1','Prof. Nikunj Patel'],
 [2,'7MP113SE26','FSA','Financial Statement Analysis','Finance',3,'Prof. Parag Rijwani (Div. A&B)\nProf. Pankaj Agrawal (Div. C)','T6\nT5','parag@nirmauni.ac.in,\npankaj.agrawal@nirmauni.ac.in,','2\n1','Prof. Pankaj Agarwal'],
 [3,'7MP121SE26','MFS','Management of Financial Service','Finance',3,'Prof. Rajesh Shah (VF) (Div. A&B)','T5','rajeshshah@nirmauni.ac.in,',2,'Prof. Pankaj Agarwal'],
 [4,'7MP101SE26','BM','Bank Management* (With IMBA)','Finance',3,'Prof. Lalit Arora (Div. A, B & C)','T6','lalit.arora@nirmauni.ac.in,',3,'Prof. Lalit Arora'],
 [5,'7MP322SE26','S&DM','Sales & Distribution Management ( With HRM)','Marketing',3,'Prof. Pradeep Kautish (Div. A)\nProf. Abhishek Shrivastav (VF) (Div. B&C)','T4\nT4','pradeep.kautish@nirmauni.ac.in,\nabhishek.shrivastav_vf@nirmauni.ac.in,','1\n2','Prof. Pradeep Kautish'],
 [6,'7MP312SE26','I&PM','Innovation and Product Management','Marketing',3,'Prof. Sandip Trada  (Div. A) \nProf. Rupam Deb (Div. B,C)','T4\nT5','sandip@nirmauni.ac.in,\nrupam.deb@nirmauni.ac.in,','1\n2','Prof. Sandip Trada'],
 [7,'7MP325SE26','SBM','Strategic Brand Management','Marketing',3,'Prof. Sanjay Jain (Div. A & B)\nProf. Riddhi Ambavale (VF) (Div. C)','T6\nT6','sanjayjain@nirmauni.ac.in,\nriddhi.ambavale_vf@nirmauni.ac.in,','2\n1','Prof. Sanjay Jain'],
 [8,'7MP303SE26','CB','Consumer Behaviour','Marketing',3,'Prof. Himanshu Chauhan (Div. A & B)\nProf. Jayesh Aagja (Div. C)','T4\nT5','himanshuchauhan@nirmauni.ac.in,\njayeshaagja@nirmauni.ac.in','2\n1','Prof. Jayesh Aagja'],
 [9,'7MP319SE26','RMKT','Retail Marketing * (1 Division with IMBA)','Marketing',3,'Prof. Sapna Prashar (Div. A ,B)','E6','sapna@nirmauni.ac.in','2\nwith IMBA','Prof. Sapna Parashar'],
 [10,'7MP506NE26','PML','People Management and Leadership * (With IMBA)','OB',3,'Prof. Deepa Sanghavi (VF)','E6','deepasanghavi@nirmauni.ac.in,\nnidhi.bansal@nirmauni.ac.in',1,'Prof. Nidhi Bansal'],
 [11,'7MP704SE26','BI','Business Intelligence','DnA',3,'Prof. Somayya Madakam (Div. A&B)','T6','somayya.madakam@nirmauni.ac.in',2,'Prof. Somayya. Madakam'],
 [12,'7MP715SE26','MBC','Managing Business on Cloud','DnA',3,'Prof Shubham Goswami (Div. A & B ) ','309-F','shubham.goswami@nirmauni.ac.in',2,'Prof. Shubham Goswami'],
 [13,'7MP710SE26','ERP','Enterprise Resource Planning','DnA',3,' P. Ganesh (VF) (Div. A & B)','T4','pganesh@nirmauni.ac.in,\nanand.kumar@nirmauni.ac.in,',2,'Prof. Anand Kumar'],
 [14,'7MP215SE26','TQM','Total Quality Management * (MBA+IMBA)','OM',3,'Prof. Dinesh Panchal & Prof. A. B. Raju  \nVF  (Div. A & B)','E6','dinesh.panchal@nirmauni.ac.in,\nabraju@nirmauni.ac.in',2,'Prof. Dinesh panchal'],
 [15,'7MP805SE26','IB','International Business* (MBA+IMBA)','E & S',3,'Prof. Punit Saurabh','E6','punit@nirmauni.ac.in',1,'Prof. Punit Saurabh'],
 [16,'7MP901CC22','SIP','Summer Internship','General',6,'Prof. Jayesh Aagja (Div. A,B,C,D,E&F)',None,'jayeshaagja@nirmauni.ac.in',1,'Prof. Jayesh Aagja'],
]
EMBEDDED_GRID = [
 ['Date','Day','Session-I','Session-II','Session-III','Session-IV',None,'Session-V','Session-VI','Session-VII','Session-VIII','Session-IX','Session-X'],
 [None,None,'08.00AM to\n09.00AM','09.10AM to \n10.10AM','10.20AM to \n11.20PM','11.30AM to \n12.30PM','12:30PM-01:30PM','1.30PM to\n02.30PM','02.40PM to\n03.40PM','03.50PM to \n04.50PM','05.00PM to \n06.00PM','06.10PM to \n07.10PM','07.20PM to \n08.20PM'],
 [None,None,'Commencement of Classes - Term-IV (June 17, 2026)',None,None,None,None,None,None,None,None,None,None],
 [datetime(2026,6,22),'Monday',"IPM(B)/IPM('C)/IPM(D)/CB(A)","IPM(A)/CB(B)/CB('C)",'FSA(A)',"FSA(B)/FSA('C)",None,'BI(B)/IB','BI(A)',None,None,None,None],
 [datetime(2026,6,23),'Tuesday',None,"BM(A)/I&PM('C)",'BM(B)/I&PM(B)/I&PM(A)',"BM('C)/SDM(A)",None,'SBM(A)','TQM(B)/SBM(B)','TQM(A)',None,None,None],
 [datetime(2026,6,24),'Wednesday',None,None,'FSA(A)/RMKT(A)',"FSA(B)/FSA('C)/RMKT(B)",None,'IB',None,None,None,None,None],
 [datetime(2026,6,25),'Thursday',None,None,'BI(B)','BI(A)',None,'MFS(B)/SBM(A)','MFS(A)/SBM(B)',None,None,None,None],
 [datetime(2026,6,26),'Friday','Muharram (Holiday)',None,None,None,None,None,None,None,None,None,None],
 [datetime(2026,6,27),'Saturday',None,None,None,None,None,None,None,None,None,None,None],
 [datetime(2026,6,28),'Sunday',None,None,None,None,None,None,None,None,None,None,None],
]

if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", default="rosters")
    ap.add_argument("--output", default="data")
    args = ap.parse_args()
    counts = build(args.input, args.output)
    print("Build complete. Row counts:")
    for t,n in counts.items(): print(f"  {t:12} {n}")
    print(f"Validation: {sum(1 for s,_ in REPORT if s=='ERROR')} errors, "
          f"{sum(1 for s,_ in REPORT if s=='WARN')} warnings")
